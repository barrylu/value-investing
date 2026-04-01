#!/usr/bin/env python3

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from statistics import median

from openpyxl import load_workbook
from pypdf import PdfReader
import xlrd


BASE_DIR = Path("/Users/luzuoguan/ai/value-investing")
COMPANY_DIR = BASE_DIR / "年报" / "长江电力-600900"
METADATA_PATH = COMPANY_DIR / "metadata.json"
SOURCE_WORKBOOK = COMPANY_DIR / "长江电力-2003至2024-财务报表与附注明细.xlsx"
RECONCILED_WORKBOOK = COMPANY_DIR / "长江电力-2003至2024-财务报表与附注明细-已对账.xlsx"
EXCEL_DIR = COMPANY_DIR / "萝卜投研-财务页下载" / "excel"

KB_DIR = COMPANY_DIR / "知识库"
FULLTEXT_DIR = KB_DIR / "年报全文"
STRUCT_DIR = KB_DIR / "结构化"
README_PATH = KB_DIR / "README.md"
INDEX_MD_PATH = KB_DIR / "索引.md"
INDEX_JSON_PATH = KB_DIR / "索引.json"
VALIDATION_MD_PATH = KB_DIR / "校验报告.md"

ANNUAL_TITLE_RE = re.compile(r"((?:19|20)\d{2})年年度报告")
ANNUAL_COLUMN_RE = re.compile(r"^(\d{4})年报$")
TOP_SECTION_RE = re.compile(r"^(第[一二三四五六七八九十百]+节)\s*(.+)$")
TOC_LINE_RE = re.compile(r"^(第[一二三四五六七八九十百]+节)\s*(.+?)\s*[\.。．·…\-_—\s]+\s*(\d+)$")
PAGE_LABEL_PATTERNS = (
    re.compile(r"第\s*(\d+)\s*页"),
    re.compile(r"(\d+)\s*/\s*(\d+)"),
)

STATEMENT_EXCEL_FILES = {
    "资产负债表": "长江电力-资产负债表-全期间.xls",
    "利润表": "长江电力-利润表-全期间.xls",
    "现金流量表": "长江电力-现金流量表-全期间.xls",
}

IGNORE_EXCEL_ITEMS = {
    "",
    "-",
    "资产",
    "负债",
    "所有者权益(或股东权益)",
    "负债和所有者权益(或股东权益)",
    "现金及现金等价物",
    "每股收益",
}

ALIAS_MAP = {
    "营业总收入": ["营业收入"],
    "营业总成本": ["营业成本"],
    "应收票据及应收账款": ["应收账款"],
    "应收票据及应收账款:应收票据": ["应收票据"],
    "应收票据及应收账款:应收账款": ["应收账款"],
    "应收票据及应收账款:应收款项融资": ["应收款项融资"],
    "其他应收款(合计)": ["其他应收款"],
    "其他应收款(合计):其他应收款": ["其他应收款"],
    "其他应收款(合计):应收利息": ["应收利息"],
    "其他应收款(合计):应收股利": ["应收股利"],
    "其他应付款(合计)": ["其他应付款"],
    "其他应付款(合计):其他应付款": ["其他应付款"],
    "其他应付款(合计):应付利息": ["应付利息"],
    "其他应付款(合计):应付股利": ["应付股利"],
    "固定资产(合计)": ["固定资产"],
    "固定资产(合计):固定资产": ["固定资产"],
    "固定资产(合计):固定资产清理": ["固定资产清理"],
    "在建工程(合计)": ["在建工程"],
    "在建工程(合计):在建工程": ["在建工程"],
    "在建工程(合计):工程物资": ["工程物资"],
    "应付票据及应付账款": ["应付账款"],
    "应付票据及应付账款:应付票据": ["应付票据"],
    "应付票据及应付账款:应付账款": ["应付账款"],
    "归属于母公司所有者权益合计": ["归属于母公司股东权益", "归属于母公司所有者权益"],
    "所有者权益合计": ["所有者权益", "所有者权益合计"],
    "负债和所有者权益总计": ["负债和股东权益总计"],
    "归属于母公司所有者的净利润": ["归属于母公司股东的净利润"],
    "归属于母公司所有者的综合收益总额": ["归属于母公司股东的综合收益总额"],
    "其中:对联营企业和合营企业的投资收益": ["对联营企业和合营企业的投资收益"],
    "其中:利息收入": ["利息收入"],
    "其中:利息费用": ["利息费用"],
    "每股收益-基本": ["基本每股收益", "每股收益(元/股)"],
    "每股收益-稀释": ["稀释每股收益"],
}


@dataclass(frozen=True)
class ReportRecord:
    year: int
    title: str
    pdf_path: Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="构建长江电力年报知识库")
    parser.add_argument("--years", nargs="*", type=int, help="仅重建指定年份，例如 --years 2009 2011")
    return parser.parse_args()


def normalize_text(text: object) -> str:
    value = str(text or "")
    replacements = {
        "：": ":",
        "（": "(",
        "）": ")",
        "“": "",
        "”": "",
        "—": "-",
        "－": "-",
        " ": "",
        "\u3000": "",
        "\t": "",
        "\n": "",
        "\r": "",
        ",": "",
        "，": "",
        ".": "",
        "。": "",
        "·": "",
        "(": "",
        ")": "",
        "/": "",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    return value.strip()


def normalize_line(line: str) -> str:
    line = line.replace("\u3000", " ")
    line = re.sub(r"[ \t]+", " ", line)
    return line.strip()


def sanitize_toc_title(title: str) -> str:
    title = re.sub(r"[\.。．·…\-_—]+", " ", title).strip()
    title = re.sub(r"\s+", " ", title)
    return title


def is_toc_page(page: dict) -> bool:
    for line in page["lines"][:6]:
        compact = normalize_text(line)
        if compact == "目录" or compact.startswith("目录第"):
            return True
    return False


def count_toc_lines(page: dict) -> int:
    count = 0
    for line in page["lines"][:80]:
        normalized = re.sub(r"\s+", " ", line)
        if TOC_LINE_RE.match(normalized):
            count += 1
    return count


def looks_like_toc_page(page: dict) -> bool:
    if is_toc_page(page):
        return True
    return count_toc_lines(page) >= 2


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in {"-", "—", "－", "None"}:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def values_close(left: float | None, right: float | None, *, tolerance: float) -> bool:
    if left is None or right is None:
        return False
    return abs(left - right) <= tolerance


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def slugify_title(title: str) -> str:
    safe = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", title).strip("-")
    return safe or "annual-report"


def load_reports() -> list[ReportRecord]:
    records = json.loads(METADATA_PATH.read_text(encoding="utf-8"))
    reports: list[ReportRecord] = []
    for item in records:
        if item.get("kind") != "年报":
            continue
        title = item.get("title", "")
        matched = ANNUAL_TITLE_RE.search(title)
        if not matched:
            continue
        pdf_path = COMPANY_DIR / item["file"]
        if not pdf_path.exists():
            raise FileNotFoundError(f"缺少年报 PDF：{pdf_path}")
        reports.append(ReportRecord(year=int(matched.group(1)), title=title, pdf_path=pdf_path))
    return sorted(reports, key=lambda item: item.year)


def extract_pdf_pages(pdf_path: Path) -> list[dict]:
    logging.getLogger("pypdf").setLevel(logging.ERROR)
    reader = PdfReader(str(pdf_path))
    pages: list[dict] = []
    for idx, page in enumerate(reader.pages, start=1):
        layout_text = page.extract_text(extraction_mode="layout") or ""
        plain_text = page.extract_text(extraction_mode="plain") or ""
        layout_lines = [normalize_line(line) for line in layout_text.splitlines()]
        plain_lines = [normalize_line(line) for line in plain_text.splitlines()]
        layout_nonempty = [line for line in layout_lines if line]
        plain_nonempty = [line for line in plain_lines if line]
        use_plain = False
        if len("".join(layout_nonempty)) < 80 and len("".join(plain_nonempty)) > len("".join(layout_nonempty)):
            use_plain = True
        lines = plain_lines if use_plain else layout_lines
        nonempty_lines = [line for line in lines if line]
        compact_text = "\n".join(nonempty_lines)
        page_label = ""
        for line in nonempty_lines[:12]:
            for pattern in PAGE_LABEL_PATTERNS:
                matched = pattern.search(line)
                if matched:
                    page_label = matched.group(0)
                    break
            if page_label:
                break
        pages.append(
            {
                "pdf_page": idx,
                "page_label": page_label,
                "char_count": len(compact_text),
                "line_count": len(nonempty_lines),
                "first_line": nonempty_lines[0] if nonempty_lines else "",
                "text": compact_text,
                "lines": nonempty_lines,
            }
        )
    return pages


def parse_toc_entries(pages: list[dict]) -> list[dict]:
    toc_start = None
    for page in pages[:12]:
        if is_toc_page(page):
            toc_start = page["pdf_page"]
            break
    if toc_start is None:
        return []

    toc_entries: list[dict] = []
    for page in pages[toc_start - 1 : min(len(pages), toc_start + 5)]:
        if page["pdf_page"] != toc_start and not looks_like_toc_page(page):
            break
        page_entries: list[dict] = []
        pending_prefix = ""
        for line in page["lines"]:
            normalized = re.sub(r"\s+", " ", line)
            if pending_prefix:
                normalized = f"{pending_prefix}{normalized}"
                pending_prefix = ""
            matched = TOC_LINE_RE.match(normalized)
            if matched:
                title = sanitize_toc_title(matched.group(2).strip())
                page_entries.append(
                    {
                        "section_label": matched.group(1),
                        "title": title,
                        "printed_page": int(matched.group(3)),
                    }
                )
                continue
            partial = re.match(r"^(第[一二三四五六七八九十百]+节)\s*(.+)$", normalized)
            if partial and not re.search(r"\d+\s*$", normalized) and len(normalized) <= 80:
                pending_prefix = f"{partial.group(1)} {sanitize_toc_title(partial.group(2))} "
                continue
            if normalize_text(normalized).startswith("备查文件目录"):
                page_entries.append({"section_label": "", "title": "备查文件目录", "printed_page": None})
        if page_entries:
            toc_entries.extend(page_entries)
        elif toc_entries and page["pdf_page"] > toc_start:
            break
    deduped: list[dict] = []
    seen: set[tuple[str, str, int | None]] = set()
    last_printed_page = -1
    for item in toc_entries:
        if item["printed_page"] is not None and item["printed_page"] > len(pages) + 20:
            continue
        if item["printed_page"] is not None and item["printed_page"] < last_printed_page:
            continue
        key = (item["section_label"], item["title"], item["printed_page"])
        if key in seen:
            continue
        seen.add(key)
        if item["printed_page"] is not None:
            last_printed_page = item["printed_page"]
        deduped.append(item)
    return deduped


def detect_top_headings(pages: list[dict]) -> list[dict]:
    headings: list[dict] = []
    for page in pages:
        if is_toc_page(page):
            continue
        seen_labels: set[str] = set()
        for line in page["lines"][:120]:
            matched = TOP_SECTION_RE.match(line)
            if matched:
                title = matched.group(2).strip()
                if "..." in title or "…" in title or "................................................................" in title:
                    continue
                if len(title) > 40:
                    continue
                section_label = matched.group(1)
                if section_label in seen_labels:
                    continue
                seen_labels.add(section_label)
                headings.append(
                    {
                        "section_label": section_label,
                        "title": title,
                        "pdf_page": page["pdf_page"],
                        "line": line,
                    }
                )
    return headings


def build_toc_validation(toc_entries: list[dict], headings: list[dict]) -> dict:
    if not toc_entries or not headings:
        return {
            "offset": None,
            "matched": 0,
            "total": len([item for item in toc_entries if item["printed_page"] is not None]),
            "details": [],
        }

    heading_by_label = {item["section_label"]: item for item in headings}
    offsets: list[int] = []
    for entry in toc_entries:
        if entry["printed_page"] is None:
            continue
        heading = heading_by_label.get(entry["section_label"])
        if heading:
            offsets.append(heading["pdf_page"] - entry["printed_page"])
    offset = Counter(offsets).most_common(1)[0][0] if offsets else None

    details: list[dict] = []
    matched_count = 0
    for entry in toc_entries:
        heading = heading_by_label.get(entry["section_label"])
        actual_page = heading["pdf_page"] if heading else None
        expected_page = entry["printed_page"] + offset if (entry["printed_page"] is not None and offset is not None) else None
        status = "未校验"
        if actual_page is not None and expected_page is not None:
            if abs(actual_page - expected_page) <= 1:
                status = "命中"
                matched_count += 1
            else:
                status = "偏移异常"
        elif entry["printed_page"] is not None:
            status = "目录有项未命中"
        details.append(
            {
                "section_label": entry["section_label"],
                "title": entry["title"],
                "printed_page": entry["printed_page"],
                "expected_pdf_page": expected_page,
                "actual_pdf_page": actual_page,
                "status": status,
            }
        )

    return {
        "offset": offset,
        "matched": matched_count,
        "total": len([item for item in toc_entries if item["printed_page"] is not None]),
        "details": details,
    }


def workbook_sheet_rows(path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(cell or "") for cell in next(iterator)]
    rows: list[dict] = []
    for row in iterator:
        item = {headers[idx]: row[idx] for idx in range(len(headers))}
        rows.append(item)
    workbook.close()
    return rows


def load_directory_rows() -> dict[int, dict]:
    rows = workbook_sheet_rows(SOURCE_WORKBOOK, "目录")
    return {int(row["年份"]): row for row in rows if row.get("年份") is not None}


def load_year_sheet_rows(year: int) -> list[dict]:
    return workbook_sheet_rows(SOURCE_WORKBOOK, f"{year}主表")


def load_datayes_summary() -> dict[int, list[dict]]:
    if not RECONCILED_WORKBOOK.exists():
        return {}
    rows = workbook_sheet_rows(RECONCILED_WORKBOOK, "Datayes校验汇总")
    grouped: dict[int, list[dict]] = defaultdict(list)
    for row in rows:
        year = row.get("年份")
        if year is None:
            continue
        grouped[int(year)].append(row)
    return dict(grouped)


def load_excel_annual_rows(statement_name: str) -> tuple[dict[int, int], list[list[object]]]:
    book = xlrd.open_workbook(EXCEL_DIR / STATEMENT_EXCEL_FILES[statement_name])
    sheet_name = "仅年报" if "仅年报" in book.sheet_names() else book.sheet_names()[0]
    sheet = book.sheet_by_name(sheet_name)
    header = sheet.row_values(0)
    year_cols: dict[int, int] = {}
    for idx, cell in enumerate(header):
        matched = ANNUAL_COLUMN_RE.match(str(cell).strip())
        if matched:
            year_cols[int(matched.group(1))] = idx
    rows = [sheet.row_values(row_idx) for row_idx in range(1, sheet.nrows)]
    return year_cols, rows


def xls_aliases(display_name: str) -> list[str]:
    aliases = [display_name]
    aliases.extend(ALIAS_MAP.get(display_name, []))
    stripped = display_name.strip()
    if stripped.startswith("其中:"):
        aliases.append(stripped.replace("其中:", "", 1))
    if stripped.startswith("减:"):
        aliases.append(stripped.replace("减:", "", 1))
    if stripped.startswith("加:"):
        aliases.append(stripped.replace("加:", "", 1))
    if "所有者" in stripped:
        aliases.append(stripped.replace("所有者", "股东"))
    if "股东" in stripped:
        aliases.append(stripped.replace("股东", "所有者"))
    return sorted({normalize_text(item) for item in aliases if item}, key=len, reverse=True)


def excel_row_should_skip(item_name: str, current_value: float | None, previous_value: float | None) -> bool:
    compact = normalize_text(item_name)
    if not compact:
        return True
    if item_name.strip() in IGNORE_EXCEL_ITEMS:
        return True
    if current_value is None and previous_value is None:
        return True
    if len(compact) <= 2 and current_value is None:
        return True
    return False


def match_score(pdf_row: dict, display_name: str, current_value: float | None, previous_value: float | None) -> tuple[int, dict]:
    aliases = xls_aliases(display_name)
    parent_name = normalize_text(display_name.split(":", 1)[0]) if ":" in display_name else ""
    child_name = normalize_text(display_name.split(":", 1)[1]) if ":" in display_name else ""
    display_name_norm = normalize_text(display_name)
    project_name = normalize_text(pdf_row.get("项目"))
    pdf_name = normalize_text(pdf_row.get("PDF项目"))
    child_project_mismatch = bool(child_name and project_name not in {display_name_norm, child_name})
    text_pool_values = [
        project_name,
        pdf_name,
        normalize_text(pdf_row.get("来源原始行")),
    ]
    current_pdf = parse_number(pdf_row.get("当前值"))
    previous_pdf = parse_number(pdf_row.get("上年值"))
    current_xls = current_value * 1_000_000 if current_value is not None else None
    previous_xls = previous_value * 1_000_000 if previous_value is not None else None
    tolerance_current = max(20_000.0, abs(current_xls or 0) * 1e-6)
    tolerance_previous = max(20_000.0, abs(previous_xls or 0) * 1e-6)
    current_has_target = current_xls is not None
    previous_has_target = previous_xls is not None
    current_hit = (not current_has_target) or values_close(current_pdf, current_xls, tolerance=tolerance_current)
    previous_hit = (not previous_has_target) or values_close(previous_pdf, previous_xls, tolerance=tolerance_previous)

    best = 0
    for pool in text_pool_values:
        for alias in aliases:
            if not pool or not alias:
                continue
            aggregate_parent_fallback = bool(parent_name and parent_name in pool and pool != alias)
            aggregate_alias_projection = bool(parent_name and project_name == parent_name and pool == alias and pdf_name == alias)
            ambiguous_child_alias = bool(
                child_name
                and pool == alias
                and pdf_name == alias
                and project_name not in {display_name_norm, child_name}
            )
            if pool == alias:
                if aggregate_alias_projection or ambiguous_child_alias:
                    best = max(best, 40)
                    continue
                best = max(best, 120)
            elif pool.endswith(alias):
                if aggregate_parent_fallback or child_project_mismatch:
                    best = max(best, 40)
                    continue
                best = max(best, 110)
            elif alias in pool:
                if aggregate_parent_fallback or child_project_mismatch:
                    best = max(best, 40)
                    continue
                best = max(best, 90 + min(len(alias), 20))
            elif len(pool) >= 6 and pool in alias:
                best = max(best, 60)
    if current_has_target and current_hit:
        best += 80
    if previous_has_target and previous_hit:
        best += 20
    return best, {
        "current_hit": current_hit,
        "previous_hit": previous_hit,
        "pdf_current": current_pdf,
        "pdf_previous": previous_pdf,
        "xls_current": current_xls,
        "xls_previous": previous_xls,
    }


def compare_year_with_excel(year: int, year_rows: list[dict]) -> dict:
    by_statement: dict[str, list[dict]] = defaultdict(list)
    for row in year_rows:
        statement = str(row.get("报表类别") or "").strip()
        if statement:
            by_statement[statement].append(row)

    overall = {"一致": 0, "差异": 0, "未命中": 0}
    statement_summaries: dict[str, dict] = {}
    mismatch_examples: list[dict] = []

    for statement_name in ("利润表", "资产负债表", "现金流量表"):
        pdf_rows = by_statement.get(statement_name, [])
        year_cols, excel_rows = load_excel_annual_rows(statement_name)
        current_col = year_cols.get(year)
        previous_col = year_cols.get(year - 1)
        compared = {"一致": 0, "差异": 0, "未命中": 0}
        compared_rows = 0

        for row in excel_rows:
            item_name = str(row[0]).strip() if row else ""
            current_value = parse_number(row[current_col]) if current_col is not None and current_col < len(row) else None
            previous_value = parse_number(row[previous_col]) if previous_col is not None and previous_col < len(row) else None
            if excel_row_should_skip(item_name, current_value, previous_value):
                continue

            best_row = None
            best_meta = None
            best_score = -1
            for pdf_row in pdf_rows:
                score, meta = match_score(pdf_row, item_name, current_value, previous_value)
                if score > best_score:
                    best_score = score
                    best_row = pdf_row
                    best_meta = meta

            compared_rows += 1
            if best_row is None or best_score < 80:
                compared["未命中"] += 1
                mismatch_examples.append(
                    {
                        "year": year,
                        "statement": statement_name,
                        "item": item_name,
                        "status": "未命中",
                        "xls_current": current_value,
                        "pdf_item": "",
                        "pdf_current": None,
                    }
                )
                continue

            if best_meta and best_meta["current_hit"] and best_meta["previous_hit"]:
                compared["一致"] += 1
            else:
                compared["差异"] += 1
                mismatch_examples.append(
                    {
                        "year": year,
                        "statement": statement_name,
                        "item": item_name,
                        "status": "差异",
                        "xls_current": current_value,
                        "pdf_item": best_row.get("项目") or best_row.get("PDF项目") or "",
                        "pdf_current": best_meta["pdf_current"] / 1_000_000 if best_meta and best_meta["pdf_current"] is not None else None,
                    }
                )

        statement_summaries[statement_name] = {"compared_rows": compared_rows, **compared}
        for key in overall:
            overall[key] += compared[key]

    return {
        "overall": overall,
        "by_statement": statement_summaries,
        "examples": mismatch_examples[:8],
    }


def render_fulltext_markdown(report: ReportRecord, pages: list[dict], toc_entries: list[dict], toc_validation: dict, sha256: str) -> str:
    lines = [f"# {report.year}年报全文", ""]
    lines.append(f"- 标题：{report.title}")
    lines.append(f"- PDF：`{report.pdf_path.name}`")
    lines.append(f"- 页数：{len(pages)}")
    lines.append(f"- SHA256：`{sha256}`")
    lines.append(
        f"- 目录校验：{toc_validation['matched']}/{toc_validation['total']} 节命中"
        + (f"，目录页码偏移 {toc_validation['offset']}" if toc_validation["offset"] is not None else "")
    )
    lines.append("")
    if toc_entries:
        lines.append("## 目录索引")
        lines.append("")
        for item in toc_entries:
            page_text = f"第 {item['printed_page']} 页" if item["printed_page"] is not None else "未标页"
            prefix = f"{item['section_label']} " if item["section_label"] else ""
            lines.append(f"- {prefix}{item['title']} | 目录页码：{page_text}")
        lines.append("")

    lines.append("## 页级全文")
    lines.append("")
    for page in pages:
        lines.append(f"### PDF第 {page['pdf_page']} 页")
        lines.append("")
        if page["text"]:
            lines.append(page["text"])
        else:
            lines.append("[本页未提取到可用文本]")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def build_report_summary(report: ReportRecord, pages: list[dict], toc_entries: list[dict], headings: list[dict], toc_validation: dict, directory_row: dict | None, excel_comparison: dict, datayes_summary: list[dict]) -> dict:
    low_text_pages = [page["pdf_page"] for page in pages if page["char_count"] < 20]
    page_char_counts = [page["char_count"] for page in pages]
    return {
        "year": report.year,
        "title": report.title,
        "pdf_file": report.pdf_path.name,
        "pdf_sha256": file_sha256(report.pdf_path),
        "pdf_pages": len(pages),
        "median_page_chars": round(median(page_char_counts), 2) if page_char_counts else 0,
        "low_text_pages": low_text_pages,
        "toc_entries": toc_entries,
        "top_headings": headings,
        "toc_validation": toc_validation,
        "workbook_directory": directory_row or {},
        "excel_comparison": excel_comparison,
        "datayes_summary": datayes_summary,
    }


def write_index_markdown(index_rows: list[dict]) -> None:
    lines = ["# 长江电力年报知识库索引", ""]
    lines.append("- 范围：`2003-2024` 全部年报")
    lines.append("- 解析形式：每年一份全文 Markdown + 一份结构化 JSON")
    lines.append("- 校验维度：PDF 页级完整性、目录页码回勾、`excel/*.xls` 年报列对比、2016-2024 Datayes 对账摘要")
    lines.append("")
    for item in index_rows:
        excel = item["excel_comparison"]["overall"]
        datayes_parts = []
        for row in item["datayes_summary"]:
            datayes_parts.append(f"{row['校验状态']} {int(row['条目数'])}")
        lines.append(f"## {item['year']}")
        lines.append("")
        lines.append(f"- 标题：{item['title']}")
        lines.append(f"- PDF：`{item['pdf_file']}` | 页数：{item['pdf_pages']} | SHA256：`{item['pdf_sha256'][:16]}`")
        lines.append(
            f"- 目录校验：{item['toc_validation']['matched']}/{item['toc_validation']['total']} 命中"
            + (f" | 页码偏移：{item['toc_validation']['offset']}" if item["toc_validation"]["offset"] is not None else "")
        )
        lines.append(f"- 低文本页：{len(item['low_text_pages'])} 页")
        lines.append(
            f"- Excel 对比：一致 {excel['一致']} 条，差异 {excel['差异']} 条，未命中 {excel['未命中']} 条"
        )
        if datayes_parts:
            lines.append(f"- Datayes 对账：{'；'.join(datayes_parts)}")
        if item["workbook_directory"]:
            lines.append(
                f"- 主表/附注抽取：主表 {item['workbook_directory'].get('主表行数')} 行，附注 {item['workbook_directory'].get('附注行数')} 行，状态 {item['workbook_directory'].get('状态')}"
            )
        lines.append(f"- 全文：`知识库/年报全文/{item['year']}-{slugify_title(item['title'])}.md`")
        lines.append(f"- 结构化：`知识库/结构化/{item['year']}.json`")
        lines.append("")
    INDEX_MD_PATH.write_text("\n".join(lines), encoding="utf-8")


def write_validation_markdown(index_rows: list[dict]) -> None:
    lines = ["# 长江电力年报知识库校验报告", ""]
    lines.append("- 校验一：PDF 页数与全文解析页数一致")
    lines.append("- 校验二：目录节次与正文顶层节标题回勾")
    lines.append("- 校验三：`excel/*.xls` 的 `仅年报` 列与年报主表标准化结果逐项比对")
    lines.append("- 校验四：`2016-2024` 额外引用 `长江电力-2003至2024-财务报表与附注明细-已对账.xlsx` 的 Datayes 对账摘要")
    lines.append("")
    for item in index_rows:
        excel = item["excel_comparison"]
        lines.append(f"## {item['year']}")
        lines.append("")
        lines.append(
            f"- PDF 完整性：`{item['pdf_pages']}` 页已全部解析；低文本页 `{len(item['low_text_pages'])}` 页；页字符数中位数 `{item['median_page_chars']}`。"
        )
        lines.append(
            f"- 目录回勾：命中 `{item['toc_validation']['matched']}/{item['toc_validation']['total']}`；目录页码偏移 `{item['toc_validation']['offset']}`。"
        )
        if item["workbook_directory"]:
            lines.append(
                f"- 主表/附注抽取复用校验：主表页 `{item['workbook_directory'].get('主表页命中')}`；附注起始页 `{item['workbook_directory'].get('附注起始页')}`；状态 `{item['workbook_directory'].get('状态')}`。"
            )
        lines.append(
            f"- Excel 对比汇总：一致 `{excel['overall']['一致']}` 条，差异 `{excel['overall']['差异']}` 条，未命中 `{excel['overall']['未命中']}` 条。"
        )
        for statement_name, summary in excel["by_statement"].items():
            lines.append(
                f"- {statement_name}：比较 `{summary['compared_rows']}` 条；一致 `{summary['一致']}`；差异 `{summary['差异']}`；未命中 `{summary['未命中']}`。"
            )
        for example in excel["examples"][:5]:
            lines.append(
                f"- 样例：`{example['statement']}` `{example['item']}` -> {example['status']}；Excel 当前值 `{example['xls_current']}` 百万元；PDF 匹配 `{example['pdf_item']}`；PDF 当前值 `{example['pdf_current']}` 百万元。"
            )
        if item["datayes_summary"]:
            datayes_text = "；".join(f"{row['校验状态']} {int(row['条目数'])} 项" for row in item["datayes_summary"])
            lines.append(f"- Datayes 对账摘要：{datayes_text}。")
        lines.append("")
    VALIDATION_MD_PATH.write_text("\n".join(lines), encoding="utf-8")


def write_readme(index_rows: list[dict]) -> None:
    lines = ["# 长江电力年报知识库", ""]
    lines.append("本目录用于把 `2003-2024` 全部年报沉淀为后续可直接引用的本地知识库，避免重复重解析 PDF。")
    lines.append("")
    lines.append("## 目录说明")
    lines.append("")
    lines.append("- `年报全文/`：每份年报的完整 Markdown 全文，按 PDF 页码分段。")
    lines.append("- `结构化/`：每年一份 JSON，包含 PDF 哈希、页统计、目录索引、顶层节标题、Excel 对比摘要。")
    lines.append("- `索引.md`：人工可读的总索引。")
    lines.append("- `索引.json`：程序可读的总索引。")
    lines.append("- `校验报告.md`：完整校验结果。")
    lines.append("")
    lines.append("## 总体结论")
    lines.append("")
    total_reports = len(index_rows)
    total_pdf_pages = sum(item["pdf_pages"] for item in index_rows)
    total_excel_hits = sum(item["excel_comparison"]["overall"]["一致"] for item in index_rows)
    total_excel_diff = sum(item["excel_comparison"]["overall"]["差异"] for item in index_rows)
    total_excel_miss = sum(item["excel_comparison"]["overall"]["未命中"] for item in index_rows)
    lines.append(f"- 年报数量：`{total_reports}`")
    lines.append(f"- PDF 总页数：`{total_pdf_pages}`")
    lines.append(f"- Excel 对比累计：一致 `{total_excel_hits}` 条，差异 `{total_excel_diff}` 条，未命中 `{total_excel_miss}` 条")
    lines.append("- 后续引用建议：优先读 `索引.md` 定位年度，再打开对应 `年报全文/*.md` 或 `结构化/*.json`。")
    lines.append("")
    README_PATH.write_text("\n".join(lines), encoding="utf-8")


def load_existing_struct_rows() -> dict[int, dict]:
    rows: dict[int, dict] = {}
    if not STRUCT_DIR.exists():
        return rows
    for path in STRUCT_DIR.glob("*.json"):
        try:
            item = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        year = item.get("year")
        if isinstance(year, int):
            rows[year] = item
    return rows


def main() -> None:
    args = parse_args()
    selected_years = set(args.years) if args.years else None
    KB_DIR.mkdir(parents=True, exist_ok=True)
    FULLTEXT_DIR.mkdir(parents=True, exist_ok=True)
    STRUCT_DIR.mkdir(parents=True, exist_ok=True)

    all_reports = load_reports()
    reports = [report for report in all_reports if selected_years is None or report.year in selected_years]
    if not reports:
        raise SystemExit("未找到需要处理的年报。")
    directory_rows = load_directory_rows()
    datayes_summary = load_datayes_summary()
    report_summaries = load_existing_struct_rows() if selected_years else {}

    for report in reports:
        print(f"[{report.year}] 解析 {report.pdf_path.name}")
        pages = extract_pdf_pages(report.pdf_path)
        toc_entries = parse_toc_entries(pages)
        headings = detect_top_headings(pages)
        toc_validation = build_toc_validation(toc_entries, headings)
        year_rows = load_year_sheet_rows(report.year)
        excel_comparison = compare_year_with_excel(report.year, year_rows)
        report_summary = build_report_summary(
            report=report,
            pages=pages,
            toc_entries=toc_entries,
            headings=headings,
            toc_validation=toc_validation,
            directory_row=directory_rows.get(report.year),
            excel_comparison=excel_comparison,
            datayes_summary=datayes_summary.get(report.year, []),
        )

        fulltext_path = FULLTEXT_DIR / f"{report.year}-{slugify_title(report.title)}.md"
        fulltext_path.write_text(
            render_fulltext_markdown(
                report=report,
                pages=pages,
                toc_entries=toc_entries,
                toc_validation=toc_validation,
                sha256=report_summary["pdf_sha256"],
            ),
            encoding="utf-8",
        )

        struct_path = STRUCT_DIR / f"{report.year}.json"
        struct_path.write_text(
            json.dumps(report_summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        report_summaries[report.year] = report_summary

    index_rows = sorted(report_summaries.values(), key=lambda item: item["year"])
    index_rows.sort(key=lambda item: item["year"])
    INDEX_JSON_PATH.write_text(json.dumps(index_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    write_index_markdown(index_rows)
    write_validation_markdown(index_rows)
    write_readme(index_rows)
    print(f"已生成知识库：{KB_DIR}")


if __name__ == "__main__":
    main()
